import requests
import openpyxl

from openpyxl import load_workbook
from bs4 import BeautifulSoup
from datetime import date
from datetime import datetime

datos = []

# Busqueda 
url = requests.get('https://monitordolarvenezuela.com/')
soup = BeautifulSoup(url.text, 'lxml')

estructura = soup.find_all('p', limit=6)

for item in estructura:
    datos.append(item)

arreglo1 = str(datos[2]).replace('<p><b>Monitor Dólar:</b> ', '')
Dolar_Monitor = str(arreglo1).replace(' Bs.S</p>', '')
#print("Dolar Monitor: " + Dolar_Monitor)

arreglo1 = str(datos[3]).replace('<p><b>Monitor Euro:</b> ', '')
euro = str(arreglo1).replace(' Bs.S</p>', '')
# print("Euro: " + euro)

arreglo1 = str(datos[4]).replace('<p><b>BCV:</b> ', '')
BCV = str(arreglo1).replace(' Bs.S</p>', '')
# print("BCV: " + BCV)

arreglo1 = str(datos[5]).replace('<p><b>DolarToday:</b> ', '')
Dolar_Today = str(arreglo1).replace(' Bs.S</p>', '')
# print("Dolar Today: " + Dolar_Today)

today = date.today()
now = datetime.now()

now = datetime.now()
fecha = now.strftime('%d/%m/%Y')
hora = now.strftime('%H:%M:%S')

precios = [
    ('Monitor', Dolar_Monitor, fecha, hora),
    ('Euro', euro, fecha, hora),
    ('BCV', BCV, fecha, hora),
    ('Today', Dolar_Today, fecha, hora)
]

print(precios)
print(BCV)

wb = load_workbook('precios.xlsx')
hoja = wb.active

# Crea la fila del encabezado con los títulos
fila = 2
bandera = 0
CB = hoja.cell(row = fila, column = 2) #B2
CC = hoja.cell(row = fila, column = 3) #C2
CD = hoja.cell(row = fila, column = 4) #D2
CE = hoja.cell(row = fila, column = 5) #E2

while bandera == 0:
    while CB.value == None:
        B = hoja.cell(row = fila, column = 2, value = Dolar_Monitor)
        C = hoja.cell(row = fila, column = 3, value = euro)
        D = hoja.cell(row = fila, column = 4, value = BCV)
        E = hoja.cell(row = fila, column = 5, value = Dolar_Today)
        F = hoja.cell(row = fila, column = 6, value = fecha) 
        G = hoja.cell(row = fila, column = 7, value = hora) 
        bandera = 1
    fila = fila + 1
    CB = hoja.cell(row = fila, column = 2)

wb.save('precios.xlsx')